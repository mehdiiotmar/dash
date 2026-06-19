import streamlit as st
import pandas as pd
import json, os, io, hashlib, base64
from datetime import datetime, date
import plotly.express as px
import plotly.graph_objects as go

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="WMCT – WI Dashboard",
    page_icon="⚓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
*, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
#MainMenu, footer, header { visibility: hidden; }

.top-banner {
    background: linear-gradient(135deg, #00245D 0%, #003087 55%, #1a5fa8 100%);
    border-radius: 14px; padding: 18px 26px; margin-bottom: 20px;
    display: flex; align-items: center; justify-content: space-between;
}
.banner-left h1 { color: white; font-size: 22px; font-weight: 700; margin: 0; }
.banner-left p  { color: #9dc3e8; font-size: 11px; margin: 2px 0 0; letter-spacing: 1.2px; }
.banner-right   { text-align: right; }
.banner-right .phase { background: rgba(255,255,255,0.15); color: white;
    border-radius: 20px; padding: 4px 14px; font-size: 11px; font-weight: 600; }

.kpi { background: white; border-radius: 12px; padding: 16px 18px;
       border: 1px solid #e8edf5; box-shadow: 0 2px 8px rgba(0,48,135,.07);
       text-align: center; }
.kpi .num  { font-size: 34px; font-weight: 700; margin: 4px 0 2px; line-height: 1; }
.kpi .lbl  { font-size: 10px; font-weight: 600; color: #6b7280;
             text-transform: uppercase; letter-spacing: .8px; }
.kpi .sub  { font-size: 11px; color: #9ca3af; margin-top: 3px; }

.sec { font-size: 12px; font-weight: 700; color: #003087; text-transform: uppercase;
       letter-spacing: .7px; padding-bottom: 6px; border-bottom: 2px solid #003087;
       margin: 18px 0 12px; }

/* WI subject card */
.wi-card {
    background: white; border-radius: 12px; border: 1px solid #e8edf5;
    padding: 16px 18px; margin-bottom: 10px;
    box-shadow: 0 1px 4px rgba(0,48,135,.05);
    transition: box-shadow .15s;
}
.wi-card:hover { box-shadow: 0 4px 14px rgba(0,48,135,.1); }
.wi-card.selected { border-left: 4px solid #1D9E75; background: #f0fdf4; }
.wi-card.open     { border-left: 4px solid #F59E0B; }
.wi-card.pending  { border-left: 4px solid #6366F1; }

/* Submission card */
.sub-card {
    background: #f8faff; border-radius: 10px; border: 1px solid #dbeafe;
    padding: 12px 14px; margin-bottom: 8px;
}
.sub-card.best { background: #f0fdf4; border-color: #86efac; }

/* Status pills */
.pill { display:inline-block; padding: 3px 10px; border-radius: 20px;
        font-size: 11px; font-weight: 600; }
.p-open     { background:#FEF3C7; color:#92400E; }
.p-submitted{ background:#DBEAFE; color:#1E40AF; }
.p-selected { background:#D1FAE5; color:#065F46; }
.p-closed   { background:#F3F4F6; color:#374151; }

/* Progress ring helper */
.prog-wrap { background:#E5E7EB; border-radius:6px; height:8px; overflow:hidden; }
.prog-fill  { height:8px; border-radius:6px;
              background: linear-gradient(90deg,#003087,#185FA5); }

.user-badge { background:#EFF6FF; border-radius:8px; padding:8px 12px;
              font-size:12px; color:#1E40AF; font-weight:500; }

.alert-info    { background:#EFF6FF; border-left:4px solid #3B82F6;
                 padding:10px 14px; border-radius:0 8px 8px 0; font-size:13px; color:#1E40AF; margin:8px 0; }
.alert-success { background:#F0FDF4; border-left:4px solid #22C55E;
                 padding:10px 14px; border-radius:0 8px 8px 0; font-size:13px; color:#15803D; margin:8px 0; }
.alert-warn    { background:#FFFBEB; border-left:4px solid #F59E0B;
                 padding:10px 14px; border-radius:0 8px 8px 0; font-size:13px; color:#92400E; margin:8px 0; }

.sidebar-user { background: linear-gradient(135deg,#EFF6FF,#DBEAFE);
                border-radius:10px; padding:12px; margin-bottom:14px; }
</style>
""", unsafe_allow_html=True)

# ─── Constants ────────────────────────────────────────────────────────────────
DATA_FILE  = "wmct_data.json"
PHASES     = ["Ship Planning","Planning","Operations","Gate","Yard","Reefer","Security","Berth Planning","General"]
CATOS_MODS = ["Berth Planning","Billing Operation","Delivery Reservation","DG Operation",
               "Gate","Operation Management","Reefer Operation","Security Management",
               "Ship Planning","Terminal Monitoring","Yard Define","Yard Planning"]

# ─── Helpers ──────────────────────────────────────────────────────────────────
def _hash(pw): return hashlib.sha256(pw.encode()).hexdigest()

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE) as f:
            return json.load(f)
    return _default_data()

def save_data(d):
    with open(DATA_FILE, "w") as f:
        json.dump(d, f, indent=2, default=str)

def _default_data():
    users = [
        {"username":"manager",  "password":_hash("manager123"),"role":"manager","display":"Manager"},
        {"username":"user1",    "password":_hash("pass1"),      "role":"user",   "display":"User 1"},
        {"username":"user2",    "password":_hash("pass2"),      "role":"user",   "display":"User 2"},
        {"username":"user3",    "password":_hash("pass3"),      "role":"user",   "display":"User 3"},
        {"username":"user4",    "password":_hash("pass4"),      "role":"user",   "display":"User 4"},
        {"username":"user5",    "password":_hash("pass5"),      "role":"user",   "display":"User 5"},
    ]
    wi_subjects = [
        {"id":"wi_001","title":"WI: Open Vessel","module":"Ship Planning",
         "phase":"Ship Planning","nav":"CATOS Home → Ship Planning → Data → Vessel & Voyage",
         "description":"Create and configure a new vessel in CATOS system.",
         "status":"open","created_by":"manager","created_at":"2025-01-10",
         "deadline":"","selected_submission":None,"notes":""},
        {"id":"wi_002","title":"WI: Create Scenario Vessel & Voyage","module":"Ship Planning",
         "phase":"Ship Planning","nav":"Ship Planning Home → DATA → Scenario",
         "description":"Define a planning scenario for a vessel and its voyage.",
         "status":"open","created_by":"manager","created_at":"2025-01-10",
         "deadline":"","selected_submission":None,"notes":""},
        {"id":"wi_003","title":"WI: Create Stoppage","module":"Ship Planning",
         "phase":"Planning","nav":"Ship Planning → Crane Working & Hatch Plan",
         "description":"Record and manage crane or operation stoppages.",
         "status":"open","created_by":"manager","created_at":"2025-01-11",
         "deadline":"","selected_submission":None,"notes":""},
        {"id":"wi_004","title":"WI: ROB to Restow / Restow to ROB","module":"Ship Planning",
         "phase":"Operations","nav":"Ship Planning → Plan View → Bay Plan",
         "description":"Move containers from ROB status to Restow and vice versa.",
         "status":"selected","created_by":"manager","created_at":"2025-01-12",
         "deadline":"","selected_submission":"sub_004_1","notes":"Best submission from User 2"},
        {"id":"wi_005","title":"WI: Bay Split & Crane Assignment","module":"Ship Planning",
         "phase":"Planning","nav":"Ship Planning → Crane Working → Split Mode",
         "description":"Split bays and assign cranes for loading/discharge operations.",
         "status":"selected","created_by":"manager","created_at":"2025-01-12",
         "deadline":"","selected_submission":"sub_005_2","notes":"Clear and complete"},
    ]
    submissions = [
        {"id":"sub_001_1","wi_id":"wi_001","submitted_by":"user1","display":"User 1",
         "submitted_at":"2025-01-15 09:00","file_name":"WI_OpenVessel_User1.pdf",
         "file_data":None,"score":None,"manager_comment":"","is_best":False},
        {"id":"sub_001_2","wi_id":"wi_001","submitted_by":"user2","display":"User 2",
         "submitted_at":"2025-01-15 11:30","file_name":"WI_OpenVessel_User2.pdf",
         "file_data":None,"score":None,"manager_comment":"","is_best":False},
        {"id":"sub_002_1","wi_id":"wi_002","submitted_by":"user1","display":"User 1",
         "submitted_at":"2025-01-16 10:00","file_name":"WI_Scenario_User1.pdf",
         "file_data":None,"score":None,"manager_comment":"","is_best":False},
        {"id":"sub_002_2","wi_id":"wi_002","submitted_by":"user3","display":"User 3",
         "submitted_at":"2025-01-16 14:00","file_name":"WI_Scenario_User3.pdf",
         "file_data":None,"score":None,"manager_comment":"","is_best":False},
        {"id":"sub_003_1","wi_id":"wi_003","submitted_by":"user2","display":"User 2",
         "submitted_at":"2025-01-17 09:30","file_name":"WI_Stoppage_User2.pdf",
         "file_data":None,"score":None,"manager_comment":"","is_best":False},
        {"id":"sub_004_1","wi_id":"wi_004","submitted_by":"user2","display":"User 2",
         "submitted_at":"2025-01-18 10:00","file_name":"WI_ROB_User2.pdf",
         "file_data":None,"score":4,"manager_comment":"Very well documented","is_best":True},
        {"id":"sub_004_2","wi_id":"wi_004","submitted_by":"user3","display":"User 3",
         "submitted_at":"2025-01-18 15:00","file_name":"WI_ROB_User3.pdf",
         "file_data":None,"score":3,"manager_comment":"Good but missing screenshots","is_best":False},
        {"id":"sub_005_2","wi_id":"wi_005","submitted_by":"user2","display":"User 2",
         "submitted_at":"2025-01-19 11:00","file_name":"WI_BaySplit_User2.pdf",
         "file_data":None,"score":5,"manager_comment":"Perfect - selected as official WI","is_best":True},
        {"id":"sub_005_3","wi_id":"wi_005","submitted_by":"user3","display":"User 3",
         "submitted_at":"2025-01-19 13:00","file_name":"WI_BaySplit_User3.pdf",
         "file_data":None,"score":3,"manager_comment":"","is_best":False},
    ]
    return {"users":users,"wi_subjects":wi_subjects,"submissions":submissions}

# ─── Session state ────────────────────────────────────────────────────────────
if "data"       not in st.session_state: st.session_state.data = load_data()
if "logged_in"  not in st.session_state: st.session_state.logged_in = False
if "user"       not in st.session_state: st.session_state.user = None
if "page"       not in st.session_state: st.session_state.page = "dashboard"

D = st.session_state.data

# ─── Quick data accessors ─────────────────────────────────────────────────────
def wi_submissions(wi_id):
    return [s for s in D["submissions"] if s["wi_id"] == wi_id]

def user_submission(wi_id, username):
    return next((s for s in D["submissions"]
                 if s["wi_id"] == wi_id and s["submitted_by"] == username), None)

def wi_by_id(wi_id):
    return next((w for w in D["wi_subjects"] if w["id"] == wi_id), None)

def get_kpis():
    wis = D["wi_subjects"]
    total    = len(wis)
    selected = sum(1 for w in wis if w["status"] == "selected")
    open_wi  = sum(1 for w in wis if w["status"] == "open")
    total_subs = len(D["submissions"])
    pending_review = sum(1 for w in wis
                         if w["status"] == "open" and len(wi_submissions(w["id"])) > 0)
    return dict(total=total, selected=selected, open=open_wi,
                total_subs=total_subs, pending_review=pending_review)

def status_pill(s):
    map_ = {
        "open":      ("<span class='pill p-open'>🟡 Open</span>"),
        "submitted": ("<span class='pill p-submitted'>🔵 Has Submissions</span>"),
        "selected":  ("<span class='pill p-selected'>✅ Best Selected</span>"),
        "closed":    ("<span class='pill p-closed'>⛔ Closed</span>"),
    }
    return map_.get(s, s)

# ─── Login ────────────────────────────────────────────────────────────────────
def login_page():
    c1, c2, c3 = st.columns([1, 1.1, 1])
    with c2:
        st.markdown("""
        <div style='text-align:center;padding:40px 0 24px'>
            <div style='font-size:52px'>⚓</div>
            <div style='font-size:26px;font-weight:700;color:#00245D;letter-spacing:1px'>WMCT</div>
            <div style='font-size:11px;color:#6B7280;letter-spacing:2.5px;margin-top:2px'>
                WEST MED CONTAINER TERMINAL
            </div>
            <div style='font-size:12px;color:#185FA5;margin-top:6px;font-weight:500'>
                CATOS · WI Training Dashboard
            </div>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login"):
            username = st.text_input("👤 Username", placeholder="Enter your username")
            password = st.text_input("🔑 Password", type="password", placeholder="Enter password")
            if st.form_submit_button("Sign In →", use_container_width=True, type="primary"):
                u = next((u for u in D["users"]
                          if u["username"] == username and u["password"] == _hash(password)), None)
                if u:
                    st.session_state.logged_in = True
                    st.session_state.user = u
                    st.rerun()
                else:
                    st.error("❌ Wrong username or password")

        st.markdown("""
        <div class='alert-info' style='margin-top:14px;font-size:12px'>
            <b>Demo logins:</b><br>
            Manager → <code>manager</code> / <code>manager123</code><br>
            Users → <code>user1</code>…<code>user5</code> / <code>pass1</code>…<code>pass5</code>
        </div>
        """, unsafe_allow_html=True)

# ─── Sidebar ──────────────────────────────────────────────────────────────────
def sidebar():
    u = st.session_state.user
    is_mgr = u["role"] == "manager"
    with st.sidebar:
        st.markdown("""
        <div style='text-align:center;padding:14px 0 6px'>
            <div style='font-size:30px'>⚓</div>
            <div style='font-weight:700;font-size:16px;color:#00245D'>WMCT</div>
            <div style='font-size:9px;color:#6B7280;letter-spacing:1.5px'>WI DASHBOARD</div>
        </div>""", unsafe_allow_html=True)

        st.markdown(f"""
        <div class='sidebar-user'>
            <div style='font-size:13px;font-weight:600;color:#1E40AF'>
                {"🏢" if is_mgr else "👤"} {u['display']}
            </div>
            <div style='font-size:11px;color:#6B7280;margin-top:2px'>
                {"Manager · Full access" if is_mgr else "User · Submit WIs"}
            </div>
        </div>""", unsafe_allow_html=True)

        # KPIs mini
        kpi = get_kpis()
        pct = int(kpi["selected"]/kpi["total"]*100) if kpi["total"] else 0
        st.markdown(f"""
        <div style='font-size:11px;color:#6B7280;margin-bottom:3px'>
            Overall progress: <b style='color:#003087'>{kpi['selected']}/{kpi['total']} WIs done</b>
        </div>
        <div class='prog-wrap' style='margin-bottom:14px'>
            <div class='prog-fill' style='width:{pct}%'></div>
        </div>""", unsafe_allow_html=True)

        st.markdown("**Menu**")

        if is_mgr:
            pages = [
                ("dashboard",   "📊 Dashboard"),
                ("wi_subjects", "📋 WI Subjects"),
                ("review",      "🔍 Review & Select Best"),
                ("analytics",   "📈 Analytics"),
                ("users",       "👥 Users"),
                ("export",      "📥 Export / Import"),
            ]
        else:
            pages = [
                ("dashboard",   "📊 Dashboard"),
                ("my_wis",      "📝 My Submissions"),
                ("export",      "📥 Export"),
            ]

        for key, label in pages:
            active = st.session_state.page == key
            if st.button(label, key=f"nav_{key}", use_container_width=True,
                         type="primary" if active else "secondary"):
                st.session_state.page = key
                st.rerun()

        st.markdown("---")
        if st.button("🚪 Logout", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.user = None
            st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
def page_dashboard():
    u = st.session_state.user
    is_mgr = u["role"] == "manager"
    kpi = get_kpis()
    pct = int(kpi["selected"]/kpi["total"]*100) if kpi["total"] else 0

    st.markdown(f"""
    <div class='top-banner'>
        <div class='banner-left'>
            <h1>⚓ WMCT — Work Instruction Dashboard</h1>
            <p>WEST MED CONTAINER TERMINAL · CATOS PRE-GO-LIVE · TRAINING PHASE</p>
        </div>
        <div class='banner-right'>
            <span class='phase'>🟢 Training Phase Active</span><br>
            <span style='color:#9dc3e8;font-size:11px;margin-top:4px;display:block'>
                {date.today().strftime("%d %B %Y")}
            </span>
        </div>
    </div>""", unsafe_allow_html=True)

    # ── KPI Row ──
    c1,c2,c3,c4,c5 = st.columns(5)
    kpis_data = [
        (c1, kpi["total"],          "Total WI Subjects",    "Created by manager",   "#185FA5"),
        (c2, kpi["selected"],       "✅ Best Selected",      f"{pct}% completed",    "#1D9E75"),
        (c3, kpi["open"],           "🟡 Open / Working",    "Awaiting submissions", "#F59E0B"),
        (c4, kpi["pending_review"], "🔍 Pending Review",    "Has submissions",      "#6366F1"),
        (c5, kpi["total_subs"],     "📄 Total Submissions", "From all users",       "#0EA5E9"),
    ]
    for col, val, lbl, sub, color in kpis_data:
        with col:
            st.markdown(f"""
            <div class='kpi'>
                <div class='lbl'>{lbl}</div>
                <div class='num' style='color:{color}'>{val}</div>
                <div class='sub'>{sub}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Main progress visual ──
    col_left, col_right = st.columns([3, 2])

    with col_left:
        st.markdown("<div class='sec'>📋 WI Subjects — Live Status</div>", unsafe_allow_html=True)

        for wi in D["wi_subjects"]:
            subs = wi_submissions(wi["id"])
            n_subs = len(subs)
            users_submitted = [s["display"] for s in subs]
            is_selected = wi["status"] == "selected"
            card_class = "selected" if is_selected else ("pending" if n_subs > 0 else "open")

            best_sub = next((s for s in subs if s.get("is_best")), None)
            best_label = f"<span style='color:#1D9E75;font-size:11px;font-weight:600'>🏆 Best: {best_sub['display']}</span>" if best_sub else ""

            subs_avatars = " ".join([
                f"<span style='background:#DBEAFE;color:#1E40AF;border-radius:20px;"
                f"padding:2px 8px;font-size:10px;font-weight:600'>{s['display']}</span>"
                for s in subs
            ])

            pct_bar = 100 if is_selected else (60 if n_subs > 0 else 5)
            bar_color = "#1D9E75" if is_selected else ("#6366F1" if n_subs > 0 else "#F59E0B")

            st.markdown(f"""
            <div class='wi-card {card_class}'>
                <div style='display:flex;align-items:flex-start;justify-content:space-between;gap:10px'>
                    <div style='flex:1'>
                        <div style='font-weight:600;font-size:14px;color:#111827'>{wi['title']}</div>
                        <div style='font-size:11px;color:#6B7280;margin-top:2px'>
                            📦 {wi['module']} &nbsp;|&nbsp; 🗂️ {wi['phase']}
                        </div>
                        <div style='font-size:11px;color:#9CA3AF;margin-top:2px;font-style:italic'>
                            🧭 {wi['nav'][:60]}{'…' if len(wi['nav'])>60 else ''}
                        </div>
                    </div>
                    <div style='text-align:right;min-width:140px'>
                        {status_pill(wi['status'])}
                        <div style='font-size:11px;color:#6B7280;margin-top:4px'>
                            {n_subs} submission{'s' if n_subs!=1 else ''}
                        </div>
                    </div>
                </div>
                <div style='margin-top:10px;display:flex;align-items:center;gap:8px;flex-wrap:wrap'>
                    {subs_avatars if subs_avatars else "<span style='font-size:11px;color:#9CA3AF'>No submissions yet</span>"}
                    {best_label}
                </div>
                <div class='prog-wrap' style='margin-top:10px'>
                    <div class='prog-fill' style='width:{pct_bar}%;background:{bar_color}'></div>
                </div>
            </div>""", unsafe_allow_html=True)

    with col_right:
        st.markdown("<div class='sec'>📊 Progress Overview</div>", unsafe_allow_html=True)

        # Donut chart
        labels = ["✅ Best Selected", "🔍 Has Submissions", "🟡 Open"]
        values = [
            sum(1 for w in D["wi_subjects"] if w["status"]=="selected"),
            sum(1 for w in D["wi_subjects"] if w["status"]=="open" and len(wi_submissions(w["id"]))>0),
            sum(1 for w in D["wi_subjects"] if w["status"]=="open" and len(wi_submissions(w["id"]))==0),
        ]
        colors = ["#1D9E75", "#6366F1", "#F59E0B"]
        fig = go.Figure(go.Pie(
            labels=labels, values=values, hole=.55,
            marker_colors=colors, textfont_size=12,
            hovertemplate="%{label}: %{value}<extra></extra>"
        ))
        fig.update_layout(
            margin=dict(l=0,r=0,t=0,b=0), height=200,
            showlegend=True, legend=dict(orientation="h", y=-0.15, font_size=11),
            annotations=[dict(text=f"<b>{pct}%</b>", x=0.5, y=0.5,
                              font_size=22, showarrow=False, font_color="#003087")]
        )
        st.plotly_chart(fig, use_container_width=True)

        # Submissions per user bar
        st.markdown("<div class='sec'>👥 Submissions per User</div>", unsafe_allow_html=True)
        users_non_mgr = [u for u in D["users"] if u["role"] != "manager"]
        user_sub_counts = []
        for u_item in users_non_mgr:
            best_count = sum(1 for s in D["submissions"]
                             if s["submitted_by"] == u_item["username"] and s.get("is_best"))
            total_count = sum(1 for s in D["submissions"]
                              if s["submitted_by"] == u_item["username"])
            user_sub_counts.append({
                "User": u_item["display"],
                "Total Submitted": total_count,
                "Best Selected": best_count,
            })

        df_u = pd.DataFrame(user_sub_counts)
        if not df_u.empty:
            fig2 = px.bar(df_u, x="User", y=["Total Submitted","Best Selected"],
                          barmode="group", height=200,
                          color_discrete_map={"Total Submitted":"#93C5FD","Best Selected":"#1D9E75"})
            fig2.update_layout(margin=dict(l=0,r=0,t=0,b=0), showlegend=True,
                               legend=dict(orientation="h",y=-0.25,font_size=10),
                               plot_bgcolor="white", paper_bgcolor="white",
                               xaxis=dict(showgrid=False), yaxis=dict(showgrid=True,gridcolor="#F3F4F6",dtick=1))
            st.plotly_chart(fig2, use_container_width=True)

        # Recent activity
        st.markdown("<div class='sec'>🕐 Recent Activity</div>", unsafe_allow_html=True)
        recent = sorted(D["submissions"], key=lambda x: x["submitted_at"], reverse=True)[:5]
        for s in recent:
            wi = wi_by_id(s["wi_id"])
            wi_title = wi["title"].replace("WI:","").strip() if wi else "?"
            best_tag = " 🏆" if s.get("is_best") else ""
            st.markdown(f"""
            <div style='padding:6px 0;border-bottom:1px solid #F3F4F6;font-size:12px'>
                <b>{s['display']}</b>{best_tag} submitted on
                <span style='color:#185FA5'>{wi_title[:30]}</span><br>
                <span style='color:#9CA3AF;font-size:10px'>{s['submitted_at'][:16]}</span>
            </div>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: WI SUBJECTS (Manager)
# ═══════════════════════════════════════════════════════════════════════════════
def page_wi_subjects():
    st.markdown("<div class='sec'>📋 WI Subjects Management</div>", unsafe_allow_html=True)

    tab_list, tab_add = st.tabs(["📋 All WI Subjects", "➕ Create New WI Subject"])

    with tab_list:
        # Filters
        f1, f2, f3 = st.columns([2,1,1])
        with f1: search = st.text_input("🔍 Search", placeholder="Search WI title or module…")
        with f2: mod_f  = st.selectbox("Module", ["All"]+CATOS_MODS)
        with f3: stat_f = st.selectbox("Status", ["All","open","selected"])

        wis = D["wi_subjects"]
        if search: wis = [w for w in wis if search.lower() in w["title"].lower() or search.lower() in w["module"].lower()]
        if mod_f  != "All": wis = [w for w in wis if w["module"] == mod_f]
        if stat_f != "All": wis = [w for w in wis if w["status"] == stat_f]

        st.markdown(f"**{len(wis)} WI subjects**")

        for wi in wis:
            subs = wi_submissions(wi["id"])
            is_sel = wi["status"] == "selected"
            best = next((s for s in subs if s.get("is_best")), None)

            with st.expander(
                f"{'✅' if is_sel else ('🔵' if subs else '🟡')}  {wi['title']}  "
                f"— {wi['module']}  ({len(subs)} submissions)"
            ):
                r1, r2 = st.columns([2,1])
                with r1:
                    st.markdown(f"**📦 Module:** {wi['module']}")
                    st.markdown(f"**🗂️ Phase:** {wi['phase']}")
                    st.markdown(f"**🧭 Navigation:** `{wi['nav']}`")
                    st.markdown(f"**📝 Description:** {wi['description']}")
                    st.markdown(f"**📅 Created:** {wi['created_at']}")
                    if wi.get("deadline"):
                        st.markdown(f"**⏰ Deadline:** {wi['deadline']}")
                with r2:
                    st.markdown(f"**Status:** {status_pill(wi['status'])}", unsafe_allow_html=True)
                    st.metric("Submissions", len(subs))
                    if best:
                        st.success(f"🏆 Best: {best['display']}")

                # Submissions list
                if subs:
                    st.markdown("**Submissions:**")
                    for s in subs:
                        best_tag = "🏆 **BEST SELECTED**" if s.get("is_best") else ""
                        score_tag = f"⭐ {s['score']}/5" if s.get("score") else ""
                        st.markdown(f"""
                        <div class='sub-card {"best" if s.get("is_best") else ""}'>
                            <b>{s['display']}</b> &nbsp; {best_tag} &nbsp;
                            <span style='color:#6B7280;font-size:11px'>{s['submitted_at'][:16]}</span>
                            &nbsp; {score_tag}<br>
                            <span style='font-size:11px;color:#185FA5'>📄 {s['file_name']}</span>
                            {f"<br><span style='font-size:11px;color:#6B7280;font-style:italic'>{s['manager_comment']}</span>" if s.get('manager_comment') else ""}
                        </div>""", unsafe_allow_html=True)
                else:
                    st.markdown("<div class='alert-info'>No submissions yet for this WI.</div>",
                                unsafe_allow_html=True)

                # Edit
                st.markdown("---")
                ec1, ec2, ec3 = st.columns(3)
                with ec1:
                    if st.button("✏️ Edit WI", key=f"edit_wi_{wi['id']}"):
                        st.session_state[f"show_edit_{wi['id']}"] = True
                with ec2:
                    if wi["status"] != "selected":
                        if st.button("🔍 Go to Review", key=f"rev_{wi['id']}"):
                            st.session_state.review_wi = wi["id"]
                            st.session_state.page = "review"
                            st.rerun()
                with ec3:
                    if st.button("🗑️ Delete", key=f"del_wi_{wi['id']}"):
                        st.session_state[f"confirm_del_wi_{wi['id']}"] = True

                if st.session_state.get(f"confirm_del_wi_{wi['id']}"):
                    st.warning(f"⚠️ Delete **{wi['title']}** and all its submissions?")
                    dc1, dc2 = st.columns(2)
                    with dc1:
                        if st.button("Yes, delete", key=f"yes_del_wi_{wi['id']}", type="primary"):
                            D["wi_subjects"] = [w for w in D["wi_subjects"] if w["id"] != wi["id"]]
                            D["submissions"]  = [s for s in D["submissions"]  if s["wi_id"] != wi["id"]]
                            save_data(D)
                            st.rerun()
                    with dc2:
                        if st.button("Cancel", key=f"no_del_wi_{wi['id']}"):
                            st.session_state[f"confirm_del_wi_{wi['id']}"] = False

                if st.session_state.get(f"show_edit_{wi['id']}"):
                    with st.form(f"form_edit_{wi['id']}"):
                        st.markdown("**Edit WI Subject**")
                        e1, e2 = st.columns(2)
                        with e1:
                            new_title = st.text_input("Title", value=wi["title"])
                            new_mod   = st.selectbox("Module", CATOS_MODS,
                                                     index=CATOS_MODS.index(wi["module"]) if wi["module"] in CATOS_MODS else 0)
                            new_phase = st.selectbox("Phase", PHASES,
                                                     index=PHASES.index(wi["phase"]) if wi["phase"] in PHASES else 0)
                        with e2:
                            new_nav   = st.text_input("Navigation", value=wi["nav"])
                            new_desc  = st.text_area("Description", value=wi["description"], height=80)
                            new_dl    = st.text_input("Deadline (optional)", value=wi.get("deadline",""))
                        if st.form_submit_button("💾 Save Changes", type="primary"):
                            for w in D["wi_subjects"]:
                                if w["id"] == wi["id"]:
                                    w["title"]       = new_title
                                    w["module"]      = new_mod
                                    w["phase"]       = new_phase
                                    w["nav"]         = new_nav
                                    w["description"] = new_desc
                                    w["deadline"]    = new_dl
                            save_data(D)
                            st.session_state[f"show_edit_{wi['id']}"] = False
                            st.success("✅ Updated!")
                            st.rerun()

    with tab_add:
        st.markdown("#### ➕ Create a New WI Subject")
        st.markdown("""
        <div class='alert-info'>
            After creating a WI subject, all users will see it and can submit their version.
            You then review all submissions and select the best one.
        </div>""", unsafe_allow_html=True)

        with st.form("form_add_wi"):
            a1, a2 = st.columns(2)
            with a1:
                new_title = st.text_input("WI Title *", placeholder="e.g. WI: Create Discharge List")
                new_mod   = st.selectbox("CATOS Module *", CATOS_MODS)
                new_phase = st.selectbox("Phase / Category", PHASES)
                new_dl    = st.text_input("Deadline (optional)", placeholder="2025-02-01")
            with a2:
                new_nav   = st.text_input("Navigation Path", placeholder="CATOS Home → Module → Feature")
                new_desc  = st.text_area("Description / Instructions for users",
                                          placeholder="Describe what users need to document…", height=120)

            submitted = st.form_submit_button("✅ Create WI Subject", use_container_width=True, type="primary")
            if submitted:
                if not new_title.strip():
                    st.error("WI Title is required.")
                else:
                    title = new_title.strip()
                    if not title.startswith("WI:"): title = "WI: " + title
                    new_id = f"wi_{len(D['wi_subjects'])+1:03d}_{int(datetime.now().timestamp())}"
                    D["wi_subjects"].append({
                        "id": new_id, "title": title, "module": new_mod,
                        "phase": new_phase, "nav": new_nav, "description": new_desc,
                        "status": "open", "created_by": "manager",
                        "created_at": str(date.today()), "deadline": new_dl,
                        "selected_submission": None, "notes": ""
                    })
                    save_data(D)
                    st.success(f"✅ WI Subject **{title}** created! Users can now submit their versions.")
                    st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: REVIEW & SELECT BEST (Manager)
# ═══════════════════════════════════════════════════════════════════════════════
def page_review():
    st.markdown("<div class='sec'>🔍 Review Submissions & Select Best WI</div>", unsafe_allow_html=True)

    st.markdown("""
    <div class='alert-info'>
        <b>Manager workflow:</b> Review all user submissions for each WI subject,
        score them, add comments, then select the best one as the official WI.
        Once selected, the WI counts as <b>completed</b> in the dashboard.
    </div>""", unsafe_allow_html=True)

    # WI selector
    wis_with_subs = [w for w in D["wi_subjects"] if len(wi_submissions(w["id"])) > 0]
    wis_open      = [w for w in D["wi_subjects"] if len(wi_submissions(w["id"])) == 0 and w["status"]=="open"]
    wis_done      = [w for w in D["wi_subjects"] if w["status"] == "selected"]

    t_pending, t_done, t_empty = st.tabs([
        f"🔍 Pending Review ({len(wis_with_subs)})",
        f"✅ Best Selected ({len(wis_done)})",
        f"🟡 No Submissions Yet ({len(wis_open)})",
    ])

    def render_wi_review(wi, allow_select=True):
        subs = wi_submissions(wi["id"])
        with st.expander(
            f"{'✅' if wi['status']=='selected' else '🔍'} {wi['title']} "
            f"— {len(subs)} submission(s)",
            expanded=(st.session_state.get("review_wi") == wi["id"])
        ):
            st.markdown(f"**📦 Module:** {wi['module']} | **🗂️ Phase:** {wi['phase']}")
            st.markdown(f"**🧭 Navigation:** `{wi['nav']}`")
            if wi["description"]:
                st.markdown(f"**📝 Instructions:** {wi['description']}")
            st.markdown("---")

            if not subs:
                st.markdown("<div class='alert-warn'>No submissions yet.</div>", unsafe_allow_html=True)
                return

            for s in subs:
                is_best = s.get("is_best", False)
                card_class = "best" if is_best else ""

                st.markdown(f"""
                <div class='sub-card {card_class}'>
                    <div style='display:flex;justify-content:space-between;align-items:center'>
                        <div>
                            <b style='font-size:14px'>{s['display']}</b>
                            {"&nbsp; <span style='background:#D1FAE5;color:#065F46;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700'>🏆 BEST SELECTED</span>" if is_best else ""}
                        </div>
                        <span style='color:#6B7280;font-size:11px'>{s['submitted_at'][:16]}</span>
                    </div>
                    <div style='margin-top:4px'>
                        <span style='font-size:12px;color:#185FA5'>📄 {s['file_name']}</span>
                        {"&nbsp;&nbsp;<span style='color:#F59E0B;font-size:12px'>⭐ "+str(s['score'])+"/5</span>" if s.get('score') else ""}
                    </div>
                    {f"<div style='font-size:12px;color:#374151;margin-top:4px;font-style:italic'>💬 {s['manager_comment']}</div>" if s.get('manager_comment') else ""}
                </div>""", unsafe_allow_html=True)

                # Download if file data exists
                if s.get("file_data"):
                    try:
                        file_bytes = base64.b64decode(s["file_data"])
                        st.download_button(
                            f"⬇️ Download {s['file_name']}",
                            data=file_bytes, file_name=s["file_name"],
                            key=f"dl_{s['id']}"
                        )
                    except: pass

                if allow_select:
                    sc1, sc2, sc3 = st.columns([1,2,1])
                    with sc1:
                        score_val = st.number_input(
                            "Score (1-5)", 1, 5,
                            value=int(s.get("score") or 3),
                            key=f"score_{s['id']}"
                        )
                    with sc2:
                        comment_val = st.text_input(
                            "Comment", value=s.get("manager_comment",""),
                            key=f"comment_{s['id']}",
                            placeholder="Your feedback…"
                        )
                    with sc3:
                        if st.button("💾 Save", key=f"save_score_{s['id']}"):
                            for sub in D["submissions"]:
                                if sub["id"] == s["id"]:
                                    sub["score"]           = score_val
                                    sub["manager_comment"] = comment_val
                            save_data(D)
                            st.success("Saved!")
                            st.rerun()

                    st.markdown("<br>", unsafe_allow_html=True)

            # Select best button
            if allow_select and wi["status"] != "selected":
                st.markdown("---")
                st.markdown("#### 🏆 Select the Best Submission")
                options = {s["id"]: f"{s['display']} — {s['file_name']}" for s in subs}
                chosen = st.selectbox(
                    "Choose the best submission to mark as official WI:",
                    list(options.keys()),
                    format_func=lambda x: options[x],
                    key=f"choose_best_{wi['id']}"
                )
                col_sel1, col_sel2 = st.columns([2,1])
                with col_sel1:
                    mgr_note = st.text_input(
                        "Note for selection (optional)",
                        key=f"note_best_{wi['id']}",
                        placeholder="Why is this the best?"
                    )
                with col_sel2:
                    if st.button(f"✅ Select as Best & Mark Done",
                                 key=f"select_best_{wi['id']}", type="primary"):
                        # Mark all subs is_best=False first
                        for sub in D["submissions"]:
                            if sub["wi_id"] == wi["id"]:
                                sub["is_best"] = (sub["id"] == chosen)
                        # Update WI
                        for w in D["wi_subjects"]:
                            if w["id"] == wi["id"]:
                                w["status"]               = "selected"
                                w["selected_submission"]  = chosen
                                w["notes"]                = mgr_note
                        save_data(D)
                        st.success("🏆 Best submission selected! This WI is now marked as completed.")
                        st.balloons()
                        st.rerun()

            elif wi["status"] == "selected":
                st.markdown("""
                <div class='alert-success'>
                    ✅ Best submission has been selected. This WI is marked as completed.
                </div>""", unsafe_allow_html=True)
                if st.button("🔓 Undo Selection (Reopen)", key=f"undo_{wi['id']}"):
                    for w in D["wi_subjects"]:
                        if w["id"] == wi["id"]:
                            w["status"] = "open"
                            w["selected_submission"] = None
                    for sub in D["submissions"]:
                        if sub["wi_id"] == wi["id"]:
                            sub["is_best"] = False
                    save_data(D)
                    st.rerun()

    with t_pending:
        if wis_with_subs:
            for wi in wis_with_subs:
                render_wi_review(wi, allow_select=True)
        else:
            st.markdown("<div class='alert-info'>No WIs with pending submissions.</div>",
                        unsafe_allow_html=True)

    with t_done:
        if wis_done:
            for wi in wis_done:
                render_wi_review(wi, allow_select=False)
        else:
            st.info("No WIs selected yet.")

    with t_empty:
        if wis_open:
            for wi in wis_open:
                st.markdown(f"""
                <div class='wi-card open'>
                    <b>{wi['title']}</b> — {wi['module']}<br>
                    <span style='font-size:11px;color:#9CA3AF'>Waiting for user submissions</span>
                </div>""", unsafe_allow_html=True)
        else:
            st.success("All WIs have at least one submission!")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: MY WIs (User)
# ═══════════════════════════════════════════════════════════════════════════════
def page_my_wis():
    u = st.session_state.user
    st.markdown(f"<div class='sec'>📝 My WI Submissions — {u['display']}</div>",
                unsafe_allow_html=True)

    tab_submit, tab_mine = st.tabs(["📤 Submit a WI", "📋 My Submissions"])

    with tab_submit:
        open_wis = [w for w in D["wi_subjects"] if w["status"] == "open"]

        if not open_wis:
            st.markdown("<div class='alert-info'>No open WI subjects right now. Check back later!</div>",
                        unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class='alert-info'>
                <b>{len(open_wis)} WI subject(s) are open.</b>
                Submit your version for any WI below. The manager will review all submissions
                and select the best one as the official WI.
            </div>""", unsafe_allow_html=True)

            for wi in open_wis:
                my_sub = user_submission(wi["id"], u["username"])
                has_sub = my_sub is not None

                with st.expander(
                    f"{'✅ Already submitted' if has_sub else '📤 Submit'} — {wi['title']}"
                ):
                    st.markdown(f"**📦 Module:** {wi['module']} | **🗂️ Phase:** {wi['phase']}")
                    st.markdown(f"**🧭 Navigation:** `{wi['nav']}`")
                    if wi["description"]:
                        st.markdown(f"**📝 What to document:** {wi['description']}")
                    if wi.get("deadline"):
                        st.markdown(f"**⏰ Deadline:** {wi['deadline']}")

                    if has_sub:
                        st.markdown(f"""
                        <div class='alert-success'>
                            ✅ You submitted: <b>{my_sub['file_name']}</b>
                            on {my_sub['submitted_at'][:16]}<br>
                            {'⭐ Score: '+str(my_sub['score'])+'/5' if my_sub.get('score') else ''}
                            {'<br>💬 '+my_sub['manager_comment'] if my_sub.get('manager_comment') else ''}
                        </div>""", unsafe_allow_html=True)

                        if st.button(f"🔄 Replace my submission", key=f"replace_{wi['id']}"):
                            st.session_state[f"replacing_{wi['id']}"] = True

                    if not has_sub or st.session_state.get(f"replacing_{wi['id']}"):
                        with st.form(f"submit_form_{wi['id']}"):
                            st.markdown("**Upload your WI document:**")
                            uploaded_file = st.file_uploader(
                                "Upload file (PDF, Word, Excel)",
                                type=["pdf","docx","xlsx","doc","pptx"],
                                key=f"file_{wi['id']}"
                            )
                            notes = st.text_area(
                                "Additional notes (optional)",
                                placeholder="Any notes for the manager…",
                                key=f"notes_{wi['id']}", height=80
                            )

                            if st.form_submit_button("📤 Submit My Version", type="primary",
                                                      use_container_width=True):
                                if not uploaded_file:
                                    st.error("Please upload a file.")
                                else:
                                    file_data = base64.b64encode(uploaded_file.read()).decode()
                                    sub_id = f"sub_{wi['id']}_{u['username']}_{int(datetime.now().timestamp())}"

                                    # Remove old submission if replacing
                                    if has_sub:
                                        D["submissions"] = [
                                            s for s in D["submissions"]
                                            if not (s["wi_id"]==wi["id"] and s["submitted_by"]==u["username"])
                                        ]

                                    D["submissions"].append({
                                        "id": sub_id,
                                        "wi_id": wi["id"],
                                        "submitted_by": u["username"],
                                        "display": u["display"],
                                        "submitted_at": str(datetime.now())[:19],
                                        "file_name": uploaded_file.name,
                                        "file_data": file_data,
                                        "notes": notes,
                                        "score": None,
                                        "manager_comment": "",
                                        "is_best": False
                                    })
                                    save_data(D)
                                    st.session_state[f"replacing_{wi['id']}"] = False
                                    st.success(f"✅ Submitted **{uploaded_file.name}** successfully!")
                                    st.rerun()

    with tab_mine:
        my_subs = [s for s in D["submissions"] if s["submitted_by"] == u["username"]]
        if not my_subs:
            st.info("You haven't submitted anything yet.")
        else:
            st.markdown(f"**{len(my_subs)} submission(s)**")
            for s in sorted(my_subs, key=lambda x: x["submitted_at"], reverse=True):
                wi = wi_by_id(s["wi_id"])
                is_best = s.get("is_best", False)
                wi_status = wi["status"] if wi else "?"

                st.markdown(f"""
                <div class='sub-card {"best" if is_best else ""}'>
                    <div style='display:flex;justify-content:space-between'>
                        <b style='font-size:14px'>{wi['title'] if wi else '?'}</b>
                        {"<span style='color:#1D9E75;font-weight:700'>🏆 BEST SELECTED!</span>" if is_best else ""}
                    </div>
                    <div style='font-size:12px;color:#6B7280;margin-top:3px'>
                        📄 {s['file_name']} &nbsp;|&nbsp; 📅 {s['submitted_at'][:16]}
                    </div>
                    {"<div style='font-size:12px;margin-top:4px'>⭐ Score: <b>"+str(s['score'])+"/5</b></div>" if s.get('score') else ""}
                    {"<div style='font-size:12px;color:#374151;margin-top:4px;font-style:italic'>💬 "+s['manager_comment']+"</div>" if s.get('manager_comment') else ""}
                    {"<div style='font-size:11px;color:#9CA3AF;margin-top:4px'>WI is still open — manager hasn't selected best yet</div>" if wi_status=='open' and not is_best else ""}
                </div>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: ANALYTICS (Manager)
# ═══════════════════════════════════════════════════════════════════════════════
def page_analytics():
    st.markdown("<div class='sec'>📈 Analytics</div>", unsafe_allow_html=True)

    t1, t2, t3 = st.tabs(["🏆 User Leaderboard", "📊 WI Completion", "📦 By Module"])

    with t1:
        users_non_mgr = [u for u in D["users"] if u["role"] != "manager"]
        rows = []
        for u_item in users_non_mgr:
            my_subs = [s for s in D["submissions"] if s["submitted_by"] == u_item["username"]]
            best    = [s for s in my_subs if s.get("is_best")]
            scores  = [s["score"] for s in my_subs if s.get("score")]
            avg_sc  = sum(scores)/len(scores) if scores else 0
            rows.append({
                "User": u_item["display"],
                "Submitted": len(my_subs),
                "Best Selected": len(best),
                "Avg Score": round(avg_sc, 1),
                "Success Rate": f"{int(len(best)/len(my_subs)*100)}%" if my_subs else "—"
            })
        rows.sort(key=lambda x: (x["Best Selected"], x["Avg Score"]), reverse=True)
        medals = ["🥇","🥈","🥉","4th","5th"]
        for i, r in enumerate(rows):
            bg = "#FFFBEB" if i==0 else "white"
            st.markdown(f"""
            <div style='background:{bg};border-radius:10px;border:1px solid #E5E7EB;
                        padding:12px 16px;margin-bottom:8px;display:flex;align-items:center;gap:16px'>
                <div style='font-size:22px'>{medals[i] if i<3 else medals[i]}</div>
                <div style='flex:1'>
                    <b style='font-size:14px'>{r['User']}</b><br>
                    <span style='font-size:12px;color:#6B7280'>
                        {r['Submitted']} submitted · {r['Best Selected']} best selected · avg {r['Avg Score']}⭐
                    </span>
                </div>
                <div style='text-align:right'>
                    <span style='font-size:18px;font-weight:700;color:#1D9E75'>{r['Best Selected']}</span>
                    <span style='font-size:11px;color:#9CA3AF;display:block'>best selections</span>
                </div>
            </div>""", unsafe_allow_html=True)

    with t2:
        wis = D["wi_subjects"]
        df_wi = pd.DataFrame([{
            "WI": w["title"].replace("WI:","").strip()[:30],
            "Submissions": len(wi_submissions(w["id"])),
            "Status": w["status"].capitalize()
        } for w in wis])
        fig = px.bar(df_wi, x="WI", y="Submissions",
                     color="Status",
                     color_discrete_map={"Selected":"#1D9E75","Open":"#F59E0B"},
                     height=320, title="Submissions per WI Subject")
        fig.update_layout(margin=dict(l=0,r=0,t=30,b=0),
                          plot_bgcolor="white", paper_bgcolor="white",
                          xaxis_tickangle=-20)
        st.plotly_chart(fig, use_container_width=True)

        # Overall gauge
        kpi = get_kpis()
        pct = kpi["selected"]/kpi["total"]*100 if kpi["total"] else 0
        fig2 = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=pct,
            number={"suffix":"%","font":{"size":36,"color":"#003087"}},
            gauge={
                "axis":{"range":[0,100],"tickcolor":"#6B7280"},
                "bar":{"color":"#003087"},
                "steps":[
                    {"range":[0,40],"color":"#FEE2E2"},
                    {"range":[40,70],"color":"#FEF3C7"},
                    {"range":[70,100],"color":"#D1FAE5"},
                ],
                "threshold":{"line":{"color":"#1D9E75","width":3},"value":100}
            },
            title={"text":"Overall WI Completion","font":{"size":14}}
        ))
        fig2.update_layout(height=220, margin=dict(l=20,r=20,t=30,b=0))
        st.plotly_chart(fig2, use_container_width=True)

    with t3:
        by_mod = {}
        for w in D["wi_subjects"]:
            m = w["module"]
            if m not in by_mod:
                by_mod[m] = {"total":0,"selected":0}
            by_mod[m]["total"] += 1
            if w["status"] == "selected":
                by_mod[m]["selected"] += 1

        df_mod = pd.DataFrame([
            {"Module": m, "Total WIs": v["total"], "Completed": v["selected"],
             "Open": v["total"]-v["selected"]}
            for m, v in by_mod.items()
        ])
        fig3 = px.bar(df_mod, x="Module", y=["Completed","Open"],
                      barmode="stack", height=300,
                      color_discrete_map={"Completed":"#1D9E75","Open":"#F59E0B"})
        fig3.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                           plot_bgcolor="white",paper_bgcolor="white")
        st.plotly_chart(fig3, use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: EXPORT / IMPORT
# ═══════════════════════════════════════════════════════════════════════════════
def page_export():
    u = st.session_state.user
    is_mgr = u["role"] == "manager"
    st.markdown("<div class='sec'>📥 Export & Import</div>", unsafe_allow_html=True)

    tabs = st.tabs(["📤 Export Report", "📥 Import WIs", "💾 Backup"])

    with tabs[0]:
        st.markdown("#### Export full WI progress report to Excel")
        if st.button("📊 Generate Excel Report", type="primary"):
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "WI Progress"

            # Header
            ws.merge_cells("A1:H1")
            ws["A1"] = "WMCT — WI Training Progress Report"
            ws["A1"].font = Font(bold=True,size=14,color="FFFFFF",name="Arial")
            ws["A1"].fill = PatternFill("solid",fgColor="00245D")
            ws["A1"].alignment = Alignment(horizontal="center",vertical="center")
            ws.row_dimensions[1].height = 28

            hdrs = ["WI Title","Module","Phase","Status","Submissions","Best Selected By","Score","Created"]
            for i,h in enumerate(hdrs):
                c = ws.cell(2,i+1,h)
                c.font = Font(bold=True,color="FFFFFF",name="Arial",size=9)
                c.fill = PatternFill("solid",fgColor="003087")
                c.alignment = Alignment(horizontal="center")

            col_w = [35,18,14,14,12,18,8,12]
            for i,w in enumerate(col_w):
                ws.column_dimensions[get_column_letter(i+1)].width = w

            for j,wi in enumerate(D["wi_subjects"]):
                r = j+3
                subs = wi_submissions(wi["id"])
                best = next((s for s in subs if s.get("is_best")), None)
                best_score = best["score"] if best and best.get("score") else ""
                best_user  = best["display"] if best else ""

                ws.cell(r,1,wi["title"]).font = Font(bold=True,size=9,name="Arial")
                ws.cell(r,2,wi["module"])
                ws.cell(r,3,wi["phase"])
                status_label = "✅ Best Selected" if wi["status"]=="selected" else f"🟡 Open ({len(subs)} subs)"
                ws.cell(r,4,status_label)
                ws.cell(r,5,len(subs))
                ws.cell(r,6,best_user)
                ws.cell(r,7,best_score)
                ws.cell(r,8,wi["created_at"])

                fc = "D1FAE5" if wi["status"]=="selected" else "FEF3C7"
                for col in range(1,9):
                    ws.cell(r,col).fill = PatternFill("solid",fgColor=fc if col==4 else ("F9FAFB" if j%2==0 else "FFFFFF"))
                    ws.cell(r,col).alignment = Alignment(horizontal="center" if col>1 else "left",vertical="center")
                ws.row_dimensions[r].height = 16

            # Submissions sheet
            ws2 = wb.create_sheet("All Submissions")
            sub_hdrs = ["WI Title","Submitted By","File Name","Submitted At","Score","Comment","Best?"]
            for i,h in enumerate(sub_hdrs):
                c = ws2.cell(1,i+1,h)
                c.font = Font(bold=True,color="FFFFFF",name="Arial",size=9)
                c.fill = PatternFill("solid",fgColor="185FA5")
                c.alignment = Alignment(horizontal="center")

            for j,s in enumerate(sorted(D["submissions"],key=lambda x:x["submitted_at"],reverse=True)):
                r = j+2
                wi = wi_by_id(s["wi_id"])
                ws2.cell(r,1,wi["title"] if wi else "?")
                ws2.cell(r,2,s["display"])
                ws2.cell(r,3,s["file_name"])
                ws2.cell(r,4,s["submitted_at"][:16])
                ws2.cell(r,5,s.get("score",""))
                ws2.cell(r,6,s.get("manager_comment",""))
                ws2.cell(r,7,"🏆 YES" if s.get("is_best") else "")
                if s.get("is_best"):
                    for col in range(1,8):
                        ws2.cell(r,col).fill = PatternFill("solid",fgColor="D1FAE5")
                ws2.row_dimensions[r].height = 16

            for col in ["A","B","C","D","E","F","G"]:
                ws2.column_dimensions[col].width = 28 if col in ["A","C"] else 16

            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)
            st.download_button(
                "⬇️ Download Excel Report",
                data=buf,
                file_name=f"WMCT_WI_Report_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    with tabs[1]:
        if not is_mgr:
            st.warning("Only manager can import WI subjects.")
        else:
            st.markdown("#### Import WI Subjects from Excel")
            st.markdown("""
            <div class='alert-info'>
                Required columns: <b>title, module, phase, nav, description</b>
            </div>""", unsafe_allow_html=True)

            # Template download
            tpl = pd.DataFrame([{
                "title":"WI: Example WI","module":"Ship Planning",
                "phase":"Planning","nav":"CATOS Home → Module",
                "description":"Describe what to document here"
            }])
            buf_tpl = io.BytesIO(); tpl.to_excel(buf_tpl, index=False); buf_tpl.seek(0)
            st.download_button("⬇️ Download Template", data=buf_tpl,
                               file_name="WI_Import_Template.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            uploaded = st.file_uploader("Upload Excel file", type=["xlsx","xls"])
            if uploaded:
                try:
                    df = pd.read_excel(uploaded)
                    st.dataframe(df.head(), use_container_width=True)
                    required = {"title","module","phase"}
                    if required.issubset(set(df.columns)):
                        if st.button("📥 Import WI Subjects", type="primary"):
                            added = 0
                            for _, row in df.iterrows():
                                if pd.isna(row.get("title")): continue
                                t = str(row["title"]).strip()
                                if not t.startswith("WI:"): t = "WI: " + t
                                new_id = f"wi_imp_{int(datetime.now().timestamp())}_{added}"
                                D["wi_subjects"].append({
                                    "id":new_id,"title":t,
                                    "module":str(row.get("module","General")),
                                    "phase":str(row.get("phase","General")),
                                    "nav":str(row.get("nav","")) if pd.notna(row.get("nav")) else "",
                                    "description":str(row.get("description","")) if pd.notna(row.get("description")) else "",
                                    "status":"open","created_by":"manager",
                                    "created_at":str(date.today()),
                                    "deadline":"","selected_submission":None,"notes":""
                                })
                                added += 1
                            save_data(D)
                            st.success(f"✅ Imported {added} WI subjects!")
                            st.rerun()
                    else:
                        st.error(f"Missing required columns. Found: {list(df.columns)}")
                except Exception as e:
                    st.error(f"Error: {e}")

    with tabs[2]:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**📤 Export full backup**")
            st.download_button("⬇️ Download JSON Backup",
                               data=json.dumps(D, indent=2, default=str),
                               file_name=f"wmct_backup_{date.today()}.json",
                               mime="application/json", use_container_width=True)
        with c2:
            st.markdown("**📥 Restore from backup**")
            jf = st.file_uploader("Upload JSON backup", type=["json"])
            if jf and st.button("Restore", type="primary"):
                try:
                    restored = json.load(jf)
                    with open(DATA_FILE,"w") as f: json.dump(restored,f,indent=2)
                    st.session_state.data = restored
                    st.success("Restored!"); st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: USERS (Manager)
# ═══════════════════════════════════════════════════════════════════════════════
def page_users():
    st.markdown("<div class='sec'>👥 User Management</div>", unsafe_allow_html=True)

    t1, t2 = st.tabs(["👥 Users", "➕ Add User"])

    with t1:
        for u_item in D["users"]:
            role_icon = "🏢" if u_item["role"] == "manager" else "👤"
            sub_count = sum(1 for s in D["submissions"] if s["submitted_by"] == u_item["username"])
            best_count= sum(1 for s in D["submissions"] if s["submitted_by"] == u_item["username"] and s.get("is_best"))

            with st.expander(f"{role_icon} {u_item['display']} (@{u_item['username']}) — {sub_count} subs, {best_count} best"):
                with st.form(f"edit_u_{u_item['username']}"):
                    eu1, eu2 = st.columns(2)
                    with eu1:
                        new_disp = st.text_input("Display name", value=u_item["display"])
                        new_role = st.selectbox("Role", ["manager","user"],
                                                index=0 if u_item["role"]=="manager" else 1)
                    with eu2:
                        new_pw = st.text_input("New password (blank = no change)", type="password")
                    if st.form_submit_button("Update"):
                        u_item["display"] = new_disp
                        u_item["role"]    = new_role
                        if new_pw: u_item["password"] = _hash(new_pw)
                        # Update display name in submissions too
                        for s in D["submissions"]:
                            if s["submitted_by"] == u_item["username"]:
                                s["display"] = new_disp
                        save_data(D)
                        st.success("Updated!"); st.rerun()

    with t2:
        with st.form("add_user_form"):
            nu1, nu2 = st.columns(2)
            with nu1:
                new_uname = st.text_input("Username *")
                new_disp2 = st.text_input("Display name *")
            with nu2:
                new_pw2   = st.text_input("Password *", type="password")
                new_role2 = st.selectbox("Role", ["user","manager"])
            if st.form_submit_button("➕ Add User", type="primary"):
                if new_uname and new_pw2 and new_disp2:
                    if any(u["username"]==new_uname for u in D["users"]):
                        st.error("Username already exists.")
                    else:
                        D["users"].append({
                            "username":new_uname,"password":_hash(new_pw2),
                            "role":new_role2,"display":new_disp2
                        })
                        save_data(D)
                        st.success(f"User {new_disp2} added!")
                        st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN ROUTER
# ═══════════════════════════════════════════════════════════════════════════════
if not st.session_state.logged_in:
    login_page()
else:
    sidebar()
    u    = st.session_state.user
    page = st.session_state.page
    is_mgr = u["role"] == "manager"

    if page == "dashboard":
        page_dashboard()
    elif page == "wi_subjects" and is_mgr:
        page_wi_subjects()
    elif page == "review" and is_mgr:
        page_review()
    elif page == "analytics" and is_mgr:
        page_analytics()
    elif page == "users" and is_mgr:
        page_users()
    elif page == "export":
        page_export()
    elif page == "my_wis" and not is_mgr:
        page_my_wis()
    else:
        st.warning("Page not available.")
